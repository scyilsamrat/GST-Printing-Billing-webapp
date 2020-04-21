import openpyxl
import os
import datetime
date = datetime.datetime.now().date()
today = date.today()
d2 = str(today.strftime("%d-%m-%Y"))
from openpyxl import Workbook
xlsx = openpyxl.load_workbook('D:\Store Management\MonthlyReport\\04-2020\\04-2020.xlsx')
file_name2 = str(today.strftime("%m-%Y"))
direct1 = "D:/Store Management/MonthlyReport/" + str(file_name2) + "/"
print(date.strftime("%b-%y"))
if (os.path.isfile(direct1+file_name2+'.xlsx')):
    xlsx = openpyxl.load_workbook(direct1+file_name2+'.xlsx')
    sheet = xlsx.active
    dimensions = sheet.dimensions
    p = str(dimensions)
    print(p[-1])
else:
    wb = Workbook()
    ws=wb.active
    ws.title = "Changed Sheet"
    os.makedirs(direct1)
    wb.save(direct1+file_name2+'.xlsx')

